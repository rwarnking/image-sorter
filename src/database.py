import json
import sqlite3
from collections.abc import Iterable
from datetime import datetime
from typing import Any, Literal, Union

import xlsxwriter
from debug_messages import InfoCodes
from helper import test_time_frame, test_time_frame_outside, wbox
from openpyxl import load_workbook

PERSON_ID = 0
PERSON_NAME = 1
ARTIST_ID = 0
ARTIST_P_ID = 1
ARTIST_MAKE = 2
ARTIST_MODEL = 3
ARTIST_S_DATE = 4
ARTIST_E_DATE = 5
ARTIST_TSHIFT = 6
EVENT_ID = 0
EVENT_TITLE = 1
EVENT_S_DATE = 2
EVENT_E_DATE = 3
SEVENT_ID = 0
SEVENT_E_ID = 1
SEVENT_TITLE = 2
SEVENT_S_DATE = 3
SEVENT_E_DATE = 4
PART_ID = 0
PART_P_ID = 1
PART_E_ID = 2
PART_S_DATE = 3
PART_E_DATE = 4

COLUMN_NAMES = {
    "person": [
        "pid",
        "name",
    ],
    "artist": [
        "aid",
        "person_id",
        "make",
        "model",
        "start_date",
        "end_date",
        "time_shift",
    ],
    "participant": [
        "paid",
        "person_id",
        "event_id",
        "start_date",
        "end_date",
    ],
    "event": [
        "eid",
        "title",
        "start_date",
        "end_date",
    ],
    "subevent": [
        "seid",
        "event_id",
        "title",
        "start_date",
        "end_date",
    ],
}


class Database:
    """
    A participant: participant_id, person_id, event_id, start_date, end_date
    It was decided to link to a person instead of an artist, since like this:
    1. there can be attending persons, which are not an artist
    2. it is not necessary to have N participants for a person, that
       uses multiple (N) divices
    """

    def __init__(self, path: str = "database.db"):
        """Setup database."""
        self.conn = sqlite3.connect(
            path, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
        )
        self.conn.execute("PRAGMA foreign_keys = 1")
        self.create_tables()
        self.conn.commit()

    def close(self):
        self.conn.close()

    def create_tables(self):
        """Creates tables for events, subevents, persons, artists & participants."""
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS events( \
            eid INTEGER PRIMARY KEY ASC, \
            title STRING, start_date TIMESTAMP, end_date TIMESTAMP)"
        )

        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS subevents( \
            seid INTEGER PRIMARY KEY ASC, \
            event_id INT NOT NULL, \
            title STRING, start_date TIMESTAMP, end_date TIMESTAMP, \
            FOREIGN KEY (event_id) \
                REFERENCES events (eid) ON DELETE CASCADE ON UPDATE CASCADE)"
        )

        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS persons (pid INTEGER PRIMARY KEY ASC, name STRING)"
        )

        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS artists(\
            aid INTEGER PRIMARY KEY ASC, \
            person_id INT, make TEXT, model TEXT, \
            start_date TIMESTAMP, end_date TIMESTAMP, time_shift TEXT, \
            FOREIGN KEY (person_id) REFERENCES persons (pid) ON DELETE CASCADE ON UPDATE CASCADE)"
        )

        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS participants (\
            paid INTEGER PRIMARY KEY ASC, \
            person_id INT NOT NULL, event_id INT NOT NULL, \
            start_date TIMESTAMP, end_date TIMESTAMP, \
            FOREIGN KEY (event_id) \
                REFERENCES events (eid) ON DELETE CASCADE ON UPDATE CASCADE, \
            FOREIGN KEY (person_id) \
                REFERENCES persons (pid) ON DELETE CASCADE ON UPDATE CASCADE)"
        )

    def load_from_json(self, file: str):
        """
        Read the table data from a json file.
        If the data was not parseable the GUI will show an info message.
        """
        err = False

        with open(file) as json_file:
            data = json.load(json_file)
            for person in data["persons"]:
                err |= (
                    self.insert_person_with_id(person["pid"], person["name"])
                    == InfoCodes.ADD_ERROR
                )

            for artist in data["artists"]:
                err |= (
                    self.insert_artist_with_id(
                        artist["aid"],
                        artist["person_id"],
                        artist["make"],
                        artist["model"],
                        datetime.fromisoformat(artist["start"]["date"]),
                        datetime.fromisoformat(artist["end"]["date"]),
                        artist["timeshift"]["date"],
                    )
                    == InfoCodes.ADD_ERROR
                )

            for event in data["events"]:
                err |= (
                    self.insert_event_with_id(
                        event["eid"],
                        event["title"],
                        datetime.fromisoformat(event["start"]["date"]),
                        datetime.fromisoformat(event["end"]["date"]),
                    )
                    == InfoCodes.ADD_ERROR
                )

            for subevent in data["subevents"]:
                err |= (
                    self.insert_subevent_with_id(
                        subevent["seid"],
                        subevent["event_id"],
                        subevent["title"],
                        datetime.fromisoformat(subevent["start"]["date"]),
                        datetime.fromisoformat(subevent["end"]["date"]),
                    )
                    == InfoCodes.ADD_ERROR
                )

            for participant in data["participants"]:
                err |= (
                    self.insert_participant_with_id(
                        participant["paid"],
                        participant["person_id"],
                        participant["event_id"],
                        datetime.fromisoformat(participant["start"]["date"]),
                        datetime.fromisoformat(participant["end"]["date"]),
                    )
                    == InfoCodes.ADD_ERROR
                )

        if err:
            return InfoCodes.LOAD_SUCCESS_PARTIAL
        else:
            return InfoCodes.LOAD_SUCCESS

    def load_from_xlsx(self, file: str):
        """
        Read the table data from a json file.
        If the data was not parseable the GUI will show an info message.
        """
        err = False

        # Define variable to load the dataframe
        workbook = load_workbook(file)

        persons_df = workbook["Persons"]
        for person in persons_df.iter_rows(min_row=2):
            err |= (
                self.insert_person_with_id(
                    person[PERSON_ID].value,
                    person[PERSON_NAME].value,
                )
                == InfoCodes.ADD_ERROR
            )

        artists_df = workbook["Artists"]
        for artist in artists_df.iter_rows(min_row=2):
            err |= (
                self.insert_artist_with_id(
                    artist[ARTIST_ID].value,
                    artist[ARTIST_P_ID].value,
                    artist[ARTIST_MAKE].value,
                    artist[ARTIST_MODEL].value,
                    datetime.fromisoformat(artist[ARTIST_S_DATE].value),
                    datetime.fromisoformat(artist[ARTIST_E_DATE].value),
                    artist[ARTIST_TSHIFT].value,
                )
                == InfoCodes.ADD_ERROR
            )

        events_df = workbook["Events"]
        for event in events_df.iter_rows(min_row=2):
            err |= (
                self.insert_event_with_id(
                    event[EVENT_ID].value,
                    event[EVENT_TITLE].value,
                    datetime.fromisoformat(event[EVENT_S_DATE].value),
                    datetime.fromisoformat(event[EVENT_E_DATE].value),
                )
                == InfoCodes.ADD_ERROR
            )

        sevents_df = workbook["Subevents"]
        for subevent in sevents_df.iter_rows(min_row=2):
            err |= (
                self.insert_subevent_with_id(
                    subevent[SEVENT_ID].value,
                    subevent[SEVENT_E_ID].value,
                    subevent[SEVENT_TITLE].value,
                    datetime.fromisoformat(subevent[SEVENT_S_DATE].value),
                    datetime.fromisoformat(subevent[SEVENT_E_DATE].value),
                )
                == InfoCodes.ADD_ERROR
            )

        parts_df = workbook["Participants"]
        for participant in parts_df.iter_rows(min_row=2):
            err |= (
                self.insert_participant_with_id(
                    participant[PART_ID].value,
                    participant[PART_P_ID].value,
                    participant[PART_E_ID].value,
                    datetime.fromisoformat(participant[PART_S_DATE].value),
                    datetime.fromisoformat(participant[PART_E_DATE].value),
                )
                == InfoCodes.ADD_ERROR
            )

        if err:
            return InfoCodes.LOAD_SUCCESS_PARTIAL
        else:
            return InfoCodes.LOAD_SUCCESS

    def save_to_json(self, file: str):
        """Save all table data to a json file."""
        json_data: dict[str, list] = {
            "events": [],
            "subevents": [],
            "participants": [],
            "artists": [],
            "persons": [],
        }

        data = self.get_all("events")
        for elem in data:
            json_data["events"].append(
                {
                    "eid": elem[EVENT_ID],
                    "title": elem[EVENT_TITLE],
                    "start": {
                        "date": str(elem[EVENT_S_DATE]),
                    },
                    "end": {
                        "date": str(elem[EVENT_E_DATE]),
                    },
                }
            )

        data = self.get_all("subevents")
        for elem in data:
            json_data["subevents"].append(
                {
                    "seid": elem[SEVENT_ID],
                    "event_id": elem[SEVENT_E_ID],
                    "title": elem[SEVENT_TITLE],
                    "start": {
                        "date": str(elem[SEVENT_S_DATE]),
                    },
                    "end": {
                        "date": str(elem[SEVENT_E_DATE]),
                    },
                }
            )

        data = self.get_all("participants")
        for elem in data:
            json_data["participants"].append(
                {
                    "paid": elem[PART_ID],
                    "person_id": elem[PART_P_ID],
                    "event_id": elem[PART_E_ID],
                    "start": {
                        "date": str(elem[PART_S_DATE]),
                    },
                    "end": {
                        "date": str(elem[PART_E_DATE]),
                    },
                }
            )

        data = self.get_all("artists")
        for elem in data:
            json_data["artists"].append(
                {
                    "aid": elem[ARTIST_ID],
                    "person_id": elem[ARTIST_P_ID],
                    "make": elem[ARTIST_MAKE],
                    "model": elem[ARTIST_MODEL],
                    "start": {
                        "date": str(elem[ARTIST_S_DATE]),
                    },
                    "end": {
                        "date": str(elem[ARTIST_E_DATE]),
                    },
                    "timeshift": {
                        "date": elem[ARTIST_TSHIFT],
                    },
                }
            )

        data = self.get_all("persons")
        for elem in data:
            json_data["persons"].append(
                {
                    "pid": elem[PERSON_ID],
                    "name": elem[PERSON_NAME],
                }
            )

        with open(file, "w") as outfile:
            json.dump(json_data, outfile, indent=4)
        return InfoCodes.SAVE_SUCCESS

    def save_to_xlsx(self, file: str):
        """Save all table data to a xlsx file."""

        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(file)

        ##########
        # Events #
        ##########
        worksheet = workbook.add_worksheet("Events")

        # Write some data headers.
        worksheet.write("A1", COLUMN_NAMES["event"][EVENT_ID])
        worksheet.write("B1", COLUMN_NAMES["event"][EVENT_TITLE])
        worksheet.write("C1", COLUMN_NAMES["event"][EVENT_S_DATE])
        worksheet.write("D1", COLUMN_NAMES["event"][EVENT_E_DATE])

        row = 1
        # Iterate over the data and write it out row by row.
        data = self.get_all("events")
        for elem in data:
            worksheet.write_number(row, EVENT_ID, elem[EVENT_ID])
            worksheet.write_string(row, EVENT_TITLE, elem[EVENT_TITLE])
            worksheet.write_string(row, EVENT_S_DATE, str(elem[EVENT_S_DATE]))
            worksheet.write_string(row, EVENT_E_DATE, str(elem[EVENT_E_DATE]))
            row += 1

        #############
        # Subevents #
        #############
        worksheet = workbook.add_worksheet("Subevents")

        # Write some data headers.
        worksheet.write("A1", COLUMN_NAMES["subevent"][SEVENT_ID])
        worksheet.write("B1", COLUMN_NAMES["subevent"][SEVENT_E_ID])
        worksheet.write("C1", COLUMN_NAMES["subevent"][SEVENT_TITLE])
        worksheet.write("D1", COLUMN_NAMES["subevent"][SEVENT_S_DATE])
        worksheet.write("E1", COLUMN_NAMES["subevent"][SEVENT_E_DATE])

        row = 1
        # Iterate over the data and write it out row by row.
        data = self.get_all("subevents")
        for elem in data:
            worksheet.write_number(row, SEVENT_ID, elem[SEVENT_ID])
            worksheet.write_number(row, SEVENT_E_ID, elem[SEVENT_E_ID])
            worksheet.write_string(row, SEVENT_TITLE, elem[SEVENT_TITLE])
            worksheet.write_string(row, SEVENT_S_DATE, str(elem[SEVENT_S_DATE]))
            worksheet.write_string(row, SEVENT_E_DATE, str(elem[SEVENT_E_DATE]))
            row += 1

        ################
        # Participants #
        ################
        worksheet = workbook.add_worksheet("Participants")

        # Write some data headers.
        worksheet.write("A1", COLUMN_NAMES["participant"][PART_ID])
        worksheet.write("B1", COLUMN_NAMES["participant"][PART_P_ID])
        worksheet.write("C1", COLUMN_NAMES["participant"][PART_E_ID])
        worksheet.write("D1", COLUMN_NAMES["participant"][PART_S_DATE])
        worksheet.write("E1", COLUMN_NAMES["participant"][PART_E_DATE])

        row = 1
        # Iterate over the data and write it out row by row.
        data = self.get_all("participants")
        for elem in data:
            worksheet.write_number(row, PART_ID, elem[PART_ID])
            worksheet.write_number(row, PART_P_ID, elem[PART_P_ID])
            worksheet.write_number(row, PART_E_ID, elem[PART_E_ID])
            worksheet.write_string(row, PART_S_DATE, str(elem[PART_S_DATE]))
            worksheet.write_string(row, PART_E_DATE, str(elem[PART_E_DATE]))
            row += 1

        ###########
        # Artists #
        ###########
        worksheet = workbook.add_worksheet("Artists")

        # Write some data headers.
        worksheet.write("A1", COLUMN_NAMES["artist"][ARTIST_ID])
        worksheet.write("B1", COLUMN_NAMES["artist"][ARTIST_P_ID])
        worksheet.write("C1", COLUMN_NAMES["artist"][ARTIST_MAKE])
        worksheet.write("D1", COLUMN_NAMES["artist"][ARTIST_MODEL])
        worksheet.write("E1", COLUMN_NAMES["artist"][ARTIST_S_DATE])
        worksheet.write("F1", COLUMN_NAMES["artist"][ARTIST_E_DATE])
        worksheet.write("G1", COLUMN_NAMES["artist"][ARTIST_TSHIFT])

        row = 1
        # Iterate over the data and write it out row by row.
        data = self.get_all("artists")
        for elem in data:
            worksheet.write_number(row, ARTIST_ID, elem[ARTIST_ID])
            worksheet.write_number(row, ARTIST_P_ID, elem[ARTIST_P_ID])
            worksheet.write_string(row, ARTIST_MAKE, elem[ARTIST_MAKE])
            worksheet.write_string(row, ARTIST_MODEL, elem[ARTIST_MODEL])
            worksheet.write_string(row, ARTIST_S_DATE, str(elem[ARTIST_S_DATE]))
            worksheet.write_string(row, ARTIST_E_DATE, str(elem[ARTIST_E_DATE]))
            worksheet.write_string(row, ARTIST_TSHIFT, str(elem[ARTIST_TSHIFT]))
            row += 1

        ###########
        # Persons #
        ###########
        worksheet = workbook.add_worksheet("Persons")

        # Write some data headers.
        worksheet.write("A1", COLUMN_NAMES["person"][PERSON_ID])
        worksheet.write("B1", COLUMN_NAMES["person"][PERSON_NAME])

        row = 1
        # Iterate over the data and write it out row by row.
        data = self.get_all("persons")
        for elem in data:
            worksheet.write_number(row, PERSON_ID, elem[PERSON_ID])
            worksheet.write_string(row, PERSON_NAME, elem[PERSON_NAME])
            row += 1

        workbook.close()
        return InfoCodes.SAVE_SUCCESS

    def reorder_by_date(self):
        """
        Reorder all tables with a date attribute, such that
        the index of an element fits its date.
        This is mainly usefull for the readability of exported files.
        """
        err = False

        # Get all data in the beginning, since deleting a person also deletes the artist
        persons = self.get_all("persons")
        artists = self.get_all("artists")
        events = self.get_all("events")
        subevents = self.get_all("subevents")
        participants = self.get_all("participants")

        ###################################################
        self.clean("persons")
        persons.sort(key=lambda x: x[PERSON_NAME])
        for person in persons:
            err |= (
                self.insert_person(
                    person[PERSON_NAME],
                )
                == InfoCodes.ADD_ERROR
            )

        ###################################################
        for idx, artist in enumerate(artists):
            tmp = None
            for i, dic in enumerate(persons):
                if dic[PERSON_ID] == artist[ARTIST_P_ID]:
                    tmp = i + 1
            # TODO
            if not tmp:
                raise ValueError

            conv = list(artist)
            conv[ARTIST_P_ID] = tmp
            artists[idx] = conv

        artists.sort(key=lambda x: (x[ARTIST_P_ID], x[ARTIST_S_DATE]))
        self.clean("artists")
        for artist in artists:
            err |= (
                self.insert_artist(
                    artist[ARTIST_P_ID],
                    artist[ARTIST_MAKE],
                    artist[ARTIST_MODEL],
                    artist[ARTIST_S_DATE],
                    artist[ARTIST_E_DATE],
                    artist[ARTIST_TSHIFT],
                )
                == InfoCodes.ADD_ERROR
            )

        ###################################################
        self.clean("events")
        events.sort(key=lambda x: x[EVENT_S_DATE])
        for event in events:
            err |= (
                self.insert_event(
                    event[EVENT_TITLE],
                    event[EVENT_S_DATE],
                    event[EVENT_E_DATE],
                )
                == InfoCodes.ADD_ERROR
            )

        ###################################################
        for idx, subevent in enumerate(subevents):
            tmp = None
            for i, dic in enumerate(events):
                if dic[EVENT_ID] == subevent[SEVENT_E_ID]:
                    tmp = i + 1
            # TODO
            if not tmp:
                raise ValueError

            conv = list(subevent)
            conv[SEVENT_E_ID] = tmp
            subevents[idx] = conv

        subevents.sort(key=lambda x: (x[SEVENT_E_ID], x[SEVENT_S_DATE]))
        self.clean("subevents")
        for subevent in subevents:
            err |= (
                self.insert_subevent(
                    subevent[SEVENT_E_ID],
                    subevent[SEVENT_TITLE],
                    subevent[SEVENT_S_DATE],
                    subevent[SEVENT_E_DATE],
                )
                == InfoCodes.ADD_ERROR
            )

        ###################################################
        for idx, participant in enumerate(participants):
            tmp1 = None
            for i, dic in enumerate(persons):
                if dic[PERSON_ID] == participant[PART_P_ID]:
                    tmp1 = i + 1

            tmp2 = None
            for i, dic in enumerate(events):
                if dic[EVENT_ID] == participant[PART_E_ID]:
                    tmp2 = i + 1
            # TODO
            if not (tmp1 or tmp2):
                raise ValueError

            conv = list(participant)
            conv[ARTIST_P_ID] = tmp1
            conv[PART_E_ID] = tmp2
            participants[idx] = conv

        participants.sort(key=lambda x: (x[PART_E_ID], x[PART_S_DATE], x[PART_P_ID]))
        self.clean("participants")
        for participant in participants:
            err |= (
                self.insert_participant(
                    participant[PART_P_ID],
                    participant[PART_E_ID],
                    participant[PART_S_DATE],
                    participant[PART_E_DATE],
                )
                == InfoCodes.ADD_ERROR
            )

        if err:
            return InfoCodes.REORDER_SUCCESS_PARTIAL
        else:
            return InfoCodes.REORDER_SUCCESS

    def clean_all(self):
        """
        OBACHT: Unused
        Deletes all table contents.
        """
        self.clean("events")
        self.clean("artists")
        self.clean("persons")
        self.clean("participants")
        self.clean("subevents")

        return InfoCodes.CLEAN_ALL_SUCCESS

    ###############################################################################################
    # Generell functions
    ###############################################################################################
    def assert_tuple_list(self, list: tuple[Union[tuple[str, Any], tuple[str, Any, Any]], ...]):
        """Assertion function to guarantee that the the argument is a list of valid touples."""
        assert isinstance(list, Iterable), "Argument is not iterable!"
        assert len(list) > 0, "List is to short, no elements found!"
        for e in list:
            assert isinstance(e, tuple), f"Element of list ({e}) is not a tuple."
            assert len(e) == 2 or len(e) == 3, f"Element of list ({e}) is to short."
            assert e[0] != "", f"First part of list-element ({e}) is an empty string."
            assert isinstance(e[0], str), f"First part of list-element ({e}) is not a string."
            assert e[1] != "", f"Second part of list-element ({e}) is an empty string."
            assert e[1] is not None, f"Second part of list-element ({e}) is none."

    def get_setter_qmarks_vals(self, args: tuple[tuple[str, Any], ...]):
        """
        Convenience function to extract the setter, question marks and
        values from the given tuple (used for db querys).
        Example:
        [("key1", value1), ("key2", value2), ("key3", value3)]
        would result in
        setter: "key1, key2, key3"
        qmaks: "?, ?, ?"
        vals: (value1, value2, value3)
        """
        self.assert_tuple_list(args)
        setter: str = ""
        qmarks: str = ""
        vals: tuple[Union[str, datetime], ...] = ()
        for pair in args:
            if setter != "":
                setter += ", "
                qmarks += ", "
            setter += pair[0]
            qmarks += "?"
            vals += (pair[1],)
        return setter, qmarks, vals

    def get_where_and_vals(self, args: tuple[Union[tuple[str, Any], tuple[str, Any, Any]], ...]):
        """
        Convenience function to extract the where and values from the given tuple (for db querys).
        Example:
        [("key1", value1), ("key2", value2), ("key3", value3)]
        would result in
        where: "key1=? AND key2=? AND key3=?"
        vals: (value1, value2, value3)
        A value can also be an list, in such a case the keyword IN is used and the given
        list is split into its component values.
        """
        self.assert_tuple_list(args)
        where: str = ""
        vals: tuple[Union[str, datetime], ...] = ()
        for pair in args:
            if where != "":
                where += "AND "
            if isinstance(pair[1], list):
                where += pair[0] + " IN (" + ",".join(["?"] * len(pair[1])) + ")"
                vals += tuple(pair[1])
            else:
                where += pair[0] + "=? "
                vals += (pair[1],)

        return where, vals

    def get_setter(self, args: tuple[tuple[str, Any, Any], ...]):
        """
        Convenience function to extract the setter and values from the given tuple (for db querys).
        Example:
        [("key1", value1), ("key2", value2), ("key3", value3)]
        would result in
        setter: "key1, key2, key3"
        vals: (value1, value2, value3)
        """
        self.assert_tuple_list(args)
        setter: str = ""
        vals: tuple[Union[str, datetime], ...] = ()
        for pair in args:
            if setter != "":
                setter += ", "
            setter += pair[0] + "=? "
            vals += (pair[2],)
        return setter, vals

    def get_all(self, table: str):
        """Get all entries from the given table."""
        cur = self.conn.execute(f"SELECT * FROM {table}")
        result = cur.fetchall()
        cur.close()
        return result

    def get(self, table: str, *args: Union[tuple[str, Any], tuple[str, Any, Any]]):
        """Get all entries of the given table that match the via args given conditions."""
        # Extract the query components.
        where, vals = self.get_where_and_vals(args)
        assert where != ""
        assert len(vals) > 0

        query = f"SELECT * FROM {table} WHERE {where}"
        cur = self.conn.execute(query, vals)
        result = cur.fetchall()
        cur.close()
        return result

    def get_by_date(self, table: str, date: datetime, *args: tuple[str, Any]):
        """Get all elements from the table with the specified date."""
        where = "start_date<=? AND end_date>=?"
        vals = (date, date)
        if args:
            # Extract the query components.
            opt_where, opt_vals = self.get_where_and_vals(args)
            where += " AND " + opt_where
            vals += opt_vals

        cur = self.conn.execute(
            f"SELECT * FROM {table} WHERE {where}",
            vals,
        )
        result = cur.fetchall()
        cur.close()
        return result

    def get_by_timeframe(self, table: str, s_date: datetime, e_date: datetime):
        """Get all elements from the table with the specified date."""
        # When the elem sdate is smaller than the timeframe end date,
        # while the elem edate is bigger than the timeframe start date
        # the element is atleast partially inside the timeframe.
        where = "start_date<=? AND end_date>=?"
        vals = (e_date, s_date)

        cur = self.conn.execute(
            f"SELECT * FROM {table} WHERE {where}",
            vals,
        )
        result = cur.fetchall()
        cur.close()
        return result

    def has(
        self, table: str, *args: Union[tuple[str, Any], tuple[str, Any, Any]]
    ) -> Union[int, Literal[False]]:
        """
        This function tests the given table if it contains an element with the given attributes.
        Obviously the input must match the columns of the table.
        There are no individual table functions for checking this,
        since it is unpredictable which attributes the user might want to test against.
        For example if the function would require make and model of an artist,
        the user could not just check against the model.
        This function does allow such special cases.
        """
        result = self.get(table, *args)

        # Return the id if present
        return result[0][0] if len(result) > 0 else False

    def insert(self, table: str, *args: tuple[str, Any]):
        """Insert a new entry to the table with the attributes specified by args."""
        # Extract the query components.
        setter, qmarks, vals = self.get_setter_qmarks_vals(args)

        # If the element is not already present add
        if not self.has(table, *args):
            query = f"INSERT INTO {table} ({setter}) VALUES ({qmarks})"
            self.conn.execute(query, vals)
            self.conn.commit()
            return InfoCodes.ADD_SUCCESS
        else:
            return InfoCodes.ADD_ERROR

    def insert_ret_id(self, table: str, *args: tuple[str, Any]):
        """
        If the element is not yet present in the specified table it is added.
        In either case the id is returned.
        """
        # Extract the query components.
        setter, qmarks, vals = self.get_setter_qmarks_vals(args)

        # If the element is not already present add
        has_elem = self.has(table, *args)
        if not has_elem:
            query = f"INSERT INTO {table} ({setter}) VALUES ({qmarks})"
            cur = self.conn.execute(query, vals)
            self.conn.commit()
            return cur.lastrowid
        else:
            return has_elem

    def insert_with_id(self, table: str, *args: tuple[str, Any]):
        """
        This function adds an element to the table with the specified id.
        This should only be used, if it is important to keep the id,
        for example when the user has predefinied data which reuses the id.
        In case the user wants to load a complete database from a file,
        this would be the case, since there are already references between
        the different elements.

        Steps of the function:
        1. Check wether an identical element is already present, in which case nothing happens.
        2. Check wether the id is already taken by another element.
           In this case the present element is shifted to another id,
           and all its references must also be adjusted.
        3. Check wether the element is present using a different id.
           In this case the element is shifted to the new id,
           which is guaranteed to be unused since we would have shifted it in 2 otherwise.
           All the present references to the old id must also be shifted.
        4. Lastly the element is added.

        Currently this function is only used for loading from a file.
        Other use is discuraged.
        """
        # Extract the query components.
        setter, qmarks, vals = self.get_setter_qmarks_vals(args)

        # If the element is already present with the same id it is not added
        if self.has(table, *args):
            return InfoCodes.ADD_ERROR

        # If the id is already taken the present element is shifted out of the way.
        if has_elem := self.has(table, args[0]):
            # Shift the element to the biggest index
            # All elements with reference to the old element are also shifted via ON UPDATE CASCADE
            count = self.get_biggest_row_id(table)
            self.update(table, (args[0][0], has_elem, count + 1))

        # If the insertion element is present in the table using a different id,
        # it should be shifted to the new one. The new id is guaranteed to be available,
        # since the potential element was shifted in the previos step.
        if has_elem := self.has(table, *args[1:]):
            # All elements with reference to the old element are also shifted via ON UPDATE CASCADE
            self.update(table, (args[0][0], has_elem, args[0][1]))
            return InfoCodes.ADD_SUCCESS_AFTER_SHIFT

        # Add the element since it is not yet in the table
        query = f"INSERT INTO {table} ({setter}) VALUES ({qmarks})"
        self.conn.execute(query, vals)
        self.conn.commit()

        return InfoCodes.ADD_SUCCESS

    def get_last_row_id(self) -> int:
        """
        Returns the id of the last row added.
        https://www.alphacodingskills.com/sqlite/notes/sqlite-func-last-insert-rowid.php
        """
        # last_insert_rowid returns (0,) when no row was inserted
        query = "SELECT last_insert_rowid()"
        cursor = self.conn.execute(query)
        # Get first element of tuple
        result = cursor.fetchone()[0]
        cursor.close()

        if result > 0:
            return result
        else:
            raise IndexError("No last row for this database connection present!")

    def get_biggest_row_id(self, table: str) -> int:
        """
        Returns the biggest index available for the table.
        OBACHT: It is not possible to just use the length of the table for this,
        since the user could add 3 elements and delete one,
        then the table would be two long and the resulting index 3,
        but the needed index is 4.
        https://stackoverflow.com/questions/23742208/fetch-the-last-row-of-a-cursor-with-sqlite
        """
        query = f"SELECT * FROM {table} ORDER BY 1 DESC LIMIT 1"
        cursor = self.conn.execute(query)
        result = cursor.fetchone()
        cursor.close()

        if result:
            # Get first element of tuple
            return result[0]
        else:
            raise IndexError("No last row for this database connection present!")

    def update(self, table: str, *args: tuple[str, Any, Any]):
        """
        Update the entry of the given table defined by args.
        The args should also contain the new values for the modified entry.
        """
        # Extract the query components.
        where, vals1 = self.get_where_and_vals(args)
        setter, vals2 = self.get_setter(args)
        vals = vals2 + vals1

        if self.has(table, *args):
            query = f"UPDATE {table} SET {setter} WHERE {where}"
            self.conn.execute(query, vals)
            self.conn.commit()
            return InfoCodes.MOD_SUCCESS
        else:
            return InfoCodes.MOD_ERROR

    def delete(self, table: str, *args: tuple[str, Any]):
        """Delete the entry from the table specified by args."""
        # Extract the query components.
        where, vals = self.get_where_and_vals(args)

        if self.has(table, *args):
            query = f"DELETE FROM {table} WHERE {where}"
            self.conn.execute(query, vals)
            self.conn.commit()
            return InfoCodes.DEL_SUCCESS
        else:
            return InfoCodes.DEL_ERROR

    def clean(self, table: str):
        """Delete all entries of this database table."""
        self.conn.execute(f"DROP TABLE IF EXISTS {table}")

        # Since this is a generic function it is sadly necessary to
        # call the function to create all tables.
        # (There are no functions for creating each solo table)
        self.create_tables()

        return InfoCodes.CLEAN_SUCCESS

    def print_table(self, table: str):
        """Debugging function for printing the table contents to the console."""
        cur = self.conn.execute(f"SELECT * FROM {table}")
        res = cur.fetchall()
        cur.close()

        print(f"Table {table} is empty." if len(res) == 0 else f"Contents of table {table}:")
        for e in res:
            print(str(e))

    ###############################################################################################
    # Artist related
    ###############################################################################################
    def validate_artist(
        self,
        pid: int,
        make: str,
        model: str,
        start_date: datetime,
        end_date: datetime,
        time_shift: str,
        aid: int = 1,
    ):
        """
        Validation function to test and guarantee, that the given parameter are valid and match
        the expected input for an artist.
        """
        if not isinstance(aid, int):
            return wbox(f"Artist: A-ID {aid}, ({make}, {model}) is not an int.")
        if aid < 1:
            return wbox(f"Artist: A-ID {aid}, ({make}, {model}) smaller than 1.")
        if not isinstance(pid, int):
            return wbox(f"Artist: P-ID {pid}, ({make}, {model}) is not an int.")
        if pid < 1:
            return wbox(f"Artist: P-ID {pid}, ({make}, {model}) smaller than 1.")
        if not self.has("persons", ("pid", pid)):
            return wbox(f"Artist: Could not find person with ID {pid}, ({make}, {model}).")

        if not isinstance(make, str):
            return wbox(f"Artist: Make ({make}) is not of type string (ID: {aid}).")
        if make == "":
            return wbox(f"Artist: Make is an empty string (ID: {aid}).")

        if not isinstance(model, str):
            return wbox(f"Artist: Model ({model}) is not of type string (ID: {aid}).")
        if model == "":
            return wbox(f"Artist: Model is an empty string (ID: {aid}).")

        if not isinstance(time_shift, str):
            return wbox(f"Artist: Timeshift ({time_shift}) is not string (ID: {aid}).")
        if time_shift == "":
            return wbox(f"Artist: Timeshift is an empty string (ID: {aid}).")

        if not isinstance(start_date, datetime):
            return wbox(f"Artist: Start date ({start_date}) is not datetime (ID: {aid}).")
        if not isinstance(end_date, datetime):
            return wbox(f"Artist: End date ({end_date}) is not datetime (ID: {aid}).")
        if start_date > end_date:
            return wbox(f"Artist: End date ({end_date}) > start date ({start_date}).")

        return InfoCodes.VAL_SUCCESS

    def insert_artist(
        self,
        pid: int,
        make: str,
        model: str,
        start_date: datetime,
        end_date: datetime,
        time_shift: str,
    ):
        """Add an artist to the database table using the given data."""
        if (
            self.validate_artist(pid, make, model, start_date, end_date, time_shift)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.ADD_ERROR

        return self.insert(
            "artists",
            ("person_id", pid),
            ("make", make),
            ("model", model),
            ("start_date", start_date),
            ("end_date", end_date),
            ("time_shift", time_shift),
        )

    def insert_artist_with_id(
        self,
        aid: int,
        pid: int,
        make: str,
        model: str,
        start_date: datetime,
        end_date: datetime,
        time_shift: str,
    ):
        """
        Add an artist to the database table using the given data and a specific ID.
        Only use this function if it is import to have a specific ID.
        """
        if (
            self.validate_artist(pid, make, model, start_date, end_date, time_shift, aid)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.ADD_ERROR

        return self.insert_with_id(
            "artists",
            ("aid", aid),
            ("person_id", pid),
            ("make", make),
            ("model", model),
            ("start_date", start_date),
            ("end_date", end_date),
            ("time_shift", time_shift),
        )

    def update_artist(
        self,
        person_id: int,
        make: str,
        model: str,
        s_date: datetime,
        e_date: datetime,
        time_shift: str,
        n_person_id: int,
        n_make: str,
        n_model: str,
        n_s_date: datetime,
        n_e_date: datetime,
        n_time_shift: str,
    ):
        """
        Wrapper function to update an artist. The given arguments specify the attributes
        of the artist that shall be updated and how its new values look like.
        """
        if (
            self.validate_artist(person_id, make, model, s_date, e_date, time_shift)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.MOD_ERROR
        if (
            self.validate_artist(n_person_id, n_make, n_model, n_s_date, n_e_date, n_time_shift)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.MOD_ERROR

        # Update returns a warning in case the element was not found.
        return self.update(
            "artists",
            ("person_id", person_id, n_person_id),
            ("make", make, n_make),
            ("model", model, n_model),
            ("start_date", s_date, n_s_date),
            ("end_date", e_date, n_e_date),
            ("time_shift", time_shift, n_time_shift),
        )

    def delete_artist(
        self,
        person_id: int,
        make: str,
        model: str,
        s_date: datetime,
        e_date: datetime,
        time_shift: str,
    ):
        """Delete the specified artist."""
        if (
            self.validate_artist(person_id, make, model, s_date, e_date, time_shift)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.DEL_ERROR

        return self.delete(
            "artists",
            ("person_id", person_id),
            ("make", make),
            ("model", model),
            ("start_date", s_date),
            ("end_date", e_date),
            ("time_shift", time_shift),
        )

    def delete_artist_s(
        self, person_id: int, make: str, model: str, s_date_s: str, e_date_s: str, time_shift: str
    ):
        """Delete the specified artist using strings as dateinformation."""
        s_date = datetime.fromisoformat(s_date_s)
        e_date = datetime.fromisoformat(e_date_s)
        return self.delete_artist(person_id, make, model, s_date, e_date, time_shift)

    def test_artist_time_frame(
        self,
        artist_id: int,
        person_id: int,
        make: str,
        model: str,
        start_date: datetime,
        end_date: datetime,
    ):
        """
        Test the timeframe of an artist against the timeframes of
        each identical artist in the database.
        If this is only an artist update, the artist id is used to skip
        the timeframe of this artist which is already present in the database.

        Does not check against time_frame_swap, since this should be done beforehand.
        The time_shift is not needed, since it should only be applied to the images,
        not the time frames.
        """
        if (
            self.validate_artist(
                person_id, make, model, start_date, end_date, "_", artist_id if artist_id else 1
            )
            == InfoCodes.VAL_ERROR
        ):
            return None

        for artist in self.get(
            "artists", ("person_id", person_id), ("make", make), ("model", model)
        ):
            # Do not test against itself
            if artist_id == artist[0]:
                continue

            if err := test_time_frame(artist[4], artist[5], start_date, end_date):
                return err
        return None

    def print_artists(self):
        self.print_table("artists")

    ###############################################################################################
    # Person related
    ###############################################################################################
    def validate_person(self, name: str, pid: int = 1):
        """
        Validation function to test and guarantee, that the given parameter are valid and match
        the expected input for a person.
        """
        if not isinstance(pid, int):
            return wbox(f"Person ID ({pid}, Name: {name}) is not an int.")
        if pid < 1:
            return wbox(f"Person ID ({pid}, Name: {name}) smaller than 1.")
        if not isinstance(name, str):
            return wbox(f"Name ({name}) is not of type string (ID: {pid}).")
        if name == "":
            return wbox(f"Name is an empty string (ID: {pid}).")

        return InfoCodes.VAL_SUCCESS

    def insert_person(self, name: str):
        """Add a person to the database table using the given data."""
        if self.validate_person(name) == InfoCodes.VAL_ERROR:
            return InfoCodes.ADD_ERROR

        return self.insert("persons", ("name", name))

    def insert_person_with_id(self, pid: int, name: str):
        """
        Add a person to the database table using the given data and a specific ID.
        Only use this function if it is import to have a specific ID.
        """
        if self.validate_person(name, pid) == InfoCodes.VAL_ERROR:
            return InfoCodes.ADD_ERROR

        return self.insert_with_id("persons", ("pid", pid), ("name", name))

    def update_person(self, name: str, n_name: str):
        """
        Wrapper function to update a person. The given argument specifies the name
        of the person that shall be updated and what its new name is.
        """
        if self.validate_person(name) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR
        if self.validate_person(n_name) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR

        return self.update("persons", ("name", name, n_name))

    def get_pname(self, id: int) -> str:
        """Returns the name of the participant that has the given id."""
        res = self.get("persons", ("pid", id))
        assert len(res) > 0
        assert len(res[0]) > 0
        return res[0][PERSON_NAME]
    
    def get_pnames(self, ids: list[int]) -> str:
        """Returns the name of the participant that has the given id."""
        res = self.get("persons", ("pid", list(set(ids))))
        assert len(res) > 0
        res_dict = {}
        for person in res:
            res_dict[person[PERSON_ID]] = person[PERSON_NAME]
            assert len(person) > 0

        return [res_dict[id] for id in ids]

    def delete_person(self, name: str):
        """Delete the specified event."""
        if self.validate_person(name) == InfoCodes.VAL_ERROR:
            return InfoCodes.DEL_ERROR

        return self.delete("persons", ("name", name))

    def print_persons(self):
        self.print_table("persons")

    ###############################################################################################
    # Event related
    ###############################################################################################
    def validate_event(self, title: str, start_date: datetime, end_date: datetime, eid: int = 1):
        """
        Validation function to test and guarantee, that the given parameter are valid and match
        the expected input for an event.
        """
        if not isinstance(eid, int):
            return wbox(f"Event ID ({eid}, Title: {title}) is not an int.")
        if eid < 1:
            return wbox(f"Event ID ({eid}, Title: {title}) smaller than 1.")
        if not isinstance(title, str):
            return wbox(f"Title ({title}) is not of type string (ID: {eid}).")
        if title == "":
            return wbox(f"Title is an empty string (ID: {eid}).")
        if not isinstance(start_date, datetime):
            return wbox(f"Start date ({start_date}) is not of type datetime (ID: {eid}).")
        if not isinstance(end_date, datetime):
            return wbox(f"End date ({end_date}) is not of type datetime (ID: {eid}).")
        if start_date > end_date:
            return wbox(f"End date ({end_date}) starts before start date ({start_date}).")

        return InfoCodes.VAL_SUCCESS

    def insert_event(self, title: str, start_date: datetime, end_date: datetime):
        """Add an event to the database table using the given data."""
        if self.validate_event(title, start_date, end_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.ADD_ERROR

        return self.insert(
            "events",
            ("title", title),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def insert_event_with_id(self, eid: int, title: str, start_date: datetime, end_date: datetime):
        """
        Add an event to the database table using the given data and a specific ID.
        Only use this function if it is import to have a specific ID.
        """
        if self.validate_event(title, start_date, end_date, eid) == InfoCodes.VAL_ERROR:
            return InfoCodes.ADD_ERROR

        return self.insert_with_id(
            "events",
            ("eid", eid),
            ("title", title),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def update_event(
        self,
        title: str,
        s_date: datetime,
        e_date: datetime,
        n_title: str,
        n_s_date: datetime,
        n_e_date: datetime,
    ):
        """
        Wrapper function to update an event. The given arguments specify the attributes
        of the event that shall be updated and how its new values look like.
        """
        if self.validate_event(title, s_date, e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR
        if self.validate_event(n_title, n_s_date, n_e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR

        return self.update(
            "events",
            ("title", title, n_title),
            ("start_date", s_date, n_s_date),
            ("end_date", e_date, n_e_date),
        )

    def delete_event(self, title: str, s_date: datetime, e_date: datetime):
        """
        Delete the specified event.
        Its subevents and participants are automatically deleted too.
        """
        if self.validate_event(title, s_date, e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.DEL_ERROR

        return self.delete(
            "events", ("title", title), ("start_date", s_date), ("end_date", e_date)
        )

    def delete_event_s(self, title: str, s_date_s: str, e_date_s: str):
        """
        Delete the specified event, using strings as date information.
        Its subevents and participants are automatically deleted too.
        """
        s_date = datetime.fromisoformat(s_date_s)
        e_date = datetime.fromisoformat(e_date_s)
        return self.delete_event(title, s_date, e_date)

    def print_events(self):
        self.print_table("events")

    ###############################################################################################
    # Subevent related
    ###############################################################################################
    def validate_subevent(
        self, event_id: int, title: str, start_date: datetime, end_date: datetime, seid: int = 1
    ):
        """
        Validation function to test and guarantee, that the given parameter are valid and match
        the expected input for a subevent.
        """
        if not isinstance(seid, int):
            return wbox(f"Subevent ({seid}, {title}): SE-ID is not an int.")
        if seid < 1:
            return wbox(f"Subevent ({seid}, {title}): SE-ID smaller than 1.")
        if not isinstance(event_id, int):
            return wbox(f"Subevent ({seid}, {title}): E-ID is not an int.")
        if event_id < 1:
            return wbox(f"Subevent ({seid}, {title}): E-ID smaller than 1.")
        if not self.has("events", ("eid", event_id)):
            return wbox(f"Subevent ({seid}, {title}): Could not find event with ID.")

        if not isinstance(title, str):
            return wbox(f"Subevent ({seid}, {title}): Title is not of type string.")
        if title == "":
            return wbox(f"Subevent ({seid}, {title}): Title is an empty string.")

        if not isinstance(start_date, datetime):
            return wbox(f"Subevent ({seid}, {title}): Start date is not of type datetime.")
        if not isinstance(end_date, datetime):
            return wbox(f"Subevent ({seid}, {title}): End date is not of type datetime.")
        if start_date > end_date:
            return wbox(f"Subevent ({seid}, {title}): End date begins before start date.")

        event = self.get("events", ("eid", event_id))[0]
        if test_time_frame_outside(event[2], event[3], start_date, end_date) is not None:
            return wbox(f"Subevent ({seid}, {title}): Timeframe does not match parent event.")

        return InfoCodes.VAL_SUCCESS

    def insert_subevent(self, event_id: int, title: str, start_date: datetime, end_date: datetime):
        """Add a subevent to the database table using the given data."""
        if self.validate_subevent(event_id, title, start_date, end_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.ADD_ERROR

        return self.insert(
            "subevents",
            ("event_id", event_id),
            ("title", title),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def insert_subevent_with_id(
        self, seid: int, event_id: int, title: str, start_date: datetime, end_date: datetime
    ):
        """
        Add a subevent to the database table using the given data and a specific ID.
        Only use this function if it is import to have a specific ID.
        """
        if self.validate_subevent(event_id, title, start_date, end_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.ADD_ERROR

        return self.insert_with_id(
            "subevents",
            ("seid", seid),
            ("event_id", event_id),
            ("title", title),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def update_subevent(
        self,
        event_id: int,
        title: str,
        s_date: datetime,
        e_date: datetime,
        n_title: str,
        n_s_date: datetime,
        n_e_date: datetime,
    ):
        """
        OBACHT: unused function
        Wrapper function to update a subevent. The given arguments specify the attributes
        of the subevent that shall be updated and how its new values look like.
        """
        if self.validate_subevent(event_id, title, s_date, e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR
        if self.validate_subevent(event_id, n_title, n_s_date, n_e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR

        return self.update(
            "subevents",
            ("event_id", event_id, event_id),
            ("title", title, n_title),
            ("start_date", s_date, n_s_date),
            ("end_date", e_date, n_e_date),
        )

    def delete_subevent(self, event_id: int, title: str, s_date: datetime, e_date: datetime):
        """
        OBACHT: unused function
        Delete the specified subevent.
        """
        if self.validate_subevent(event_id, title, s_date, e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.DEL_ERROR

        return self.delete(
            "subevents",
            ("event_id", event_id),
            ("title", title),
            ("start_date", s_date),
            ("end_date", e_date),
        )

    def print_subevents(self):
        self.print_table("subevents")

    ###############################################################################################
    # Participant related
    ###############################################################################################
    def validate_participant(
        self,
        person_id: int,
        event_id: int,
        start_date: datetime,
        end_date: datetime,
        paid: int = 1,
    ):
        """
        Validation function to test and guarantee, that the given parameter are valid and match
        the expected input for a participant.
        """
        if not isinstance(paid, int):
            return wbox(f"Participant ({paid}, {person_id}): PA-ID is not an int.")
        if paid < 1:
            return wbox(f"Participant ({paid}, {person_id}): PA-ID smaller than 1.")
        if not isinstance(person_id, int):
            return wbox(f"Participant ({paid}, {person_id}): P-ID is not an int.")
        if person_id < 1:
            return wbox(f"Participant ({paid}, {person_id}): P-ID smaller than 1.")
        if not self.has("persons", ("pid", person_id)):
            return wbox(f"Participant ({paid}, {person_id}): Could not find person with ID.")
        if not isinstance(event_id, int):
            return wbox(f"Participant ({paid}, {person_id}): E-ID is not an int.")
        if event_id < 1:
            return wbox(f"Participant ({paid}, {person_id}): E-ID smaller than 1.")
        if not self.has("events", ("eid", event_id)):
            return wbox(f"Participant ({paid}, {person_id}): Could not find event with ID.")

        if not isinstance(start_date, datetime):
            return wbox(f"Participant ({paid}, {person_id}): Start date is not of type datetime.")
        if not isinstance(end_date, datetime):
            return wbox(f"Participant ({paid}, {person_id}): End date is not of type datetime.")
        if start_date > end_date:
            return wbox(f"Participant ({paid}, {person_id}): End date begins before start date.")

        event = self.get("events", ("eid", event_id))[0]
        if test_time_frame_outside(event[2], event[3], start_date, end_date) is not None:
            return wbox(f"Participant ({paid}, {person_id}): Timeframe doesnt match parent event.")

        return InfoCodes.VAL_SUCCESS

    def insert_participant(
        self, person_id: int, event_id: int, start_date: datetime, end_date: datetime
    ):
        """Add a participant to the database table using the given data."""
        if (
            self.validate_participant(person_id, event_id, start_date, end_date)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.ADD_ERROR

        return self.insert(
            "participants",
            ("person_id", person_id),
            ("event_id", event_id),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def insert_participant_with_id(
        self, paid: int, person_id: int, event_id: int, start_date: datetime, end_date: datetime
    ):
        """
        Add a participant to the database table using the given data and a specific ID.
        Only use this function if it is import to have a specific ID.
        """
        if (
            self.validate_participant(person_id, event_id, start_date, end_date, paid)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.ADD_ERROR

        return self.insert_with_id(
            "participants",
            ("paid", paid),
            ("person_id", person_id),
            ("event_id", event_id),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def update_participant(
        self,
        person_id: int,
        event_id: int,
        s_date: datetime,
        e_date: datetime,
        n_s_date: datetime,
        n_e_date: datetime,
    ):
        """
        OBACHT: unused function
        Wrapper function to update a participant. The given arguments specify the attributes
        of the participant that shall be updated and how its new values look like.
        """
        if self.validate_participant(person_id, event_id, s_date, e_date) == InfoCodes.VAL_ERROR:
            return InfoCodes.MOD_ERROR
        if (
            self.validate_participant(person_id, event_id, n_s_date, n_e_date)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.MOD_ERROR

        return self.update(
            "participants",
            ("person_id", person_id, person_id),
            ("event_id", event_id, event_id),
            ("start_date", s_date, n_s_date),
            ("end_date", e_date, n_e_date),
        )

    def delete_participant(
        self, person_id: int, event_id: int, start_date: datetime, end_date: datetime
    ):
        """
        OBACHT: unused function
        Delete the specified participant.
        """
        if (
            self.validate_participant(person_id, event_id, start_date, end_date)
            == InfoCodes.VAL_ERROR
        ):
            return InfoCodes.DEL_ERROR

        return self.delete(
            "persons",
            ("event_id", event_id),
            ("person_id", person_id),
            ("start_date", start_date),
            ("end_date", end_date),
        )

    def test_participant_time_frame(self, name: str, start_date: datetime, end_date: datetime):
        """
        Test the timeframe of the particiapnt against the timeframes of each participant
        with the same name in the database.

        Does not check against time_frame_swap, since this should be done beforehand.
        """
        if person_id := self.has("persons", ("name", name)):
            for part in self.get("participants", ("person_id", person_id)):
                if err := test_time_frame(part[3], part[4], start_date, end_date):
                    return err

            return None
        else:
            raise ValueError(f"Person with name {name} could not found!")

    def print_participants(self):
        self.print_table("participants")
